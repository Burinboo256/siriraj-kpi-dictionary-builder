from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from openpyxl import Workbook

from src.generate_template import DEFAULT_CONFIG, SHEET_NAMES, create_template_workbook
from src.utils import (
    build_index_sections,
    build_validation_issues,
    download_google_sheet_workbook,
    generate_validation_report,
    load_config,
    load_kpi_dataset,
    resolve_workbook_source,
)


class UtilsTests(unittest.TestCase):
    def test_load_kpi_dataset_returns_joined_records_from_normalized_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "template.xlsx"
            create_template_workbook(template_path)

            dataset = load_kpi_dataset(template_path)

            self.assertEqual(len(dataset.records), 3)
            self.assertEqual(dataset.records[0]["KPI_Code"], "DH0101")
            self.assertEqual(dataset.records[0]["KPI_Category"], "Cardiovascular disease (Heart disease: H)")
            self.assertEqual(dataset.records[0]["KPI_Type"], "Acute coronary syndrome (ACS)")
            self.assertIn("Numerator_Logic", dataset.records[0])
            self.assertIn("Numerator_CodeSet", dataset.records[0])

    def test_load_kpi_dataset_uses_input_form_as_authoritative_record_list(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "template.xlsx"
            create_template_workbook(template_path)

            workbook = load_workbook(template_path)
            input_sheet = workbook[SHEET_NAMES["kpi_input_form"]]
            master_sheet = workbook[SHEET_NAMES["kpi_master"]]

            input_sheet.delete_rows(3, 2)
            master_sheet.append(
                [
                    "Corporate",
                    "Standalone hidden row",
                    "ZZ9999",
                    "Hidden only",
                    "Corporate",
                    "",
                    "",
                    "Higher is better",
                    "Percent",
                    "Active",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            workbook.save(template_path)

            dataset = load_kpi_dataset(template_path)

            self.assertEqual([record["KPI_Code"] for record in dataset.records], ["DH0101"])

    def test_load_config_returns_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "template.xlsx"
            create_template_workbook(template_path)

            config = load_config(template_path)

            self.assertEqual(config["Document_Title"], "KPI Dictionary")
            self.assertEqual(config["Font_TH"], "TH Sarabun New")

    def test_load_kpi_dataset_accepts_legacy_google_sheet_shape(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "legacy.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "KPIs"
            sheet.append(
                [
                    "group_order",
                    "group_code",
                    "group_name_th",
                    "group_name_en",
                    "category_code",
                    "category_name",
                    "type_code",
                    "type_name",
                    "kpi_code",
                    "name_th",
                    "name_en",
                    "definition",
                    "objective",
                    "formula",
                    "numerator_data",
                    "denominator_data",
                    "numerator_icd",
                    "denominator_icd",
                    "frequency",
                    "benchmark",
                    "interpretation",
                    "reference",
                    "start_date",
                    "last_update_date",
                    "update_reason",
                    "notes",
                ]
            )
            sheet.append(
                [
                    1,
                    "D",
                    "กลุ่มตัวชี้วัดรายโรค",
                    "Disease",
                    "DH",
                    "Cardiovascular disease",
                    "DH01",
                    "Acute coronary syndrome",
                    "DH0101",
                    "ร้อยละการเสียชีวิตของผู้ป่วย ACS",
                    "ACS mortality",
                    "Definition TH",
                    "Objective TH",
                    "(a/b) x 100",
                    "a = deaths",
                    "b = discharges",
                    "I21.0",
                    "I21.0,I21.1",
                    "Monthly",
                    "< 9%",
                    "ค่ายิ่งน้อย = มีคุณภาพดี",
                    "THIP",
                    "1 ตุลาคม 2557",
                    "1 ตุลาคม 2564",
                    "ปรับปรุงครั้งที่ 1",
                    "หมายเหตุ",
                ]
            )
            workbook.save(workbook_path)

            dataset = load_kpi_dataset(workbook_path)

            self.assertEqual(len(dataset.records), 1)
            self.assertEqual(dataset.records[0]["KPI_Code"], "DH0101")
            self.assertEqual(dataset.records[0]["KPI_Category"], "Cardiovascular disease")
            self.assertEqual(dataset.records[0]["KPI_Type"], "Acute coronary syndrome")
            self.assertEqual(dataset.records[0]["Direction"], "Lower is better")
            self.assertEqual(dataset.config["Document_Title"], DEFAULT_CONFIG["Document_Title"])

    def test_build_index_sections_groups_by_requested_search_dimensions(self) -> None:
        records = [
            {
                "KPI_Code": "DH0101",
                "KPI_Name_TH": "A",
                "KPI_Name_EN": "A EN",
                "KPI_Category": "Cardiovascular disease",
                "KPI_Type": "Acute coronary syndrome",
                "Frequency": "Monthly",
                "Reference_Team": "ทีมโรคหัวใจ",
                "ICD_Procedure_Index": "I21.0",
            }
        ]

        sections = build_index_sections(records)

        self.assertIn("Index by KPI Code", sections)
        self.assertIn("Index by KPI Category", sections)
        self.assertEqual(sections["Index by KPI Code"][0]["KPI_Code"], "DH0101")

    def test_validation_report_flags_duplicates_missing_fields_and_bad_related_code(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "template.xlsx"
            report_path = Path(tmpdir) / "validation_report.xlsx"
            create_template_workbook(template_path)

            workbook = load_workbook(template_path)
            master_sheet = workbook[SHEET_NAMES["kpi_master"]]
            input_sheet = workbook[SHEET_NAMES["kpi_input_form"]]

            master_sheet.append([cell.value for cell in master_sheet[2]])
            master_sheet["C5"] = "DH0101"
            input_sheet["H2"] = ""
            input_sheet["M2"] = "InvalidFrequency"
            workbook.save(template_path)

            dataset = load_kpi_dataset(template_path)
            issues = build_validation_issues(dataset)
            generate_validation_report(report_path, issues)

            issue_types = {issue["Issue_Type"] for issue in issues}
            self.assertIn("Duplicate KPI_Code", issue_types)
            self.assertIn("Missing formula", issue_types)
            self.assertIn("Invalid dropdown values", issue_types)
            self.assertIn("Missing required fields", issue_types)
            self.assertTrue(report_path.exists())

    def test_resolve_workbook_source_accepts_google_sheet_url(self) -> None:
        source = "https://docs.google.com/spreadsheets/d/1s6LB0a_j4k-RhBTKAfZwQVDda25xzgmbGFDTLRC1LvM/edit?usp=sharing"
        with tempfile.TemporaryDirectory() as tmpdir:
            target = Path(tmpdir) / "downloaded.xlsx"
            with patch("src.utils.download_google_sheet_workbook", return_value=target) as mocked:
                resolved = resolve_workbook_source(source, Path(tmpdir))

            self.assertEqual(resolved, target)
            mocked.assert_called_once()

    def test_resolve_workbook_source_accepts_local_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "template.xlsx"
            create_template_workbook(template_path)

            resolved = resolve_workbook_source(str(template_path), Path(tmpdir))

            self.assertEqual(resolved, template_path)

    def test_download_google_sheet_workbook_builds_export_url_from_sheet_id(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            target_dir = Path(tmpdir)

            class FakeResponse:
                def __init__(self) -> None:
                    self.payload = b"excel-bytes"

                def read(self) -> bytes:
                    return self.payload

                def __enter__(self):
                    return self

                def __exit__(self, exc_type, exc, tb) -> None:
                    return None

            with patch("src.utils.urlopen", return_value=FakeResponse()) as mocked:
                output_path = download_google_sheet_workbook(
                    "1s6LB0a_j4k-RhBTKAfZwQVDda25xzgmbGFDTLRC1LvM",
                    target_dir,
                )

            self.assertTrue(output_path.exists())
            self.assertEqual(output_path.read_bytes(), b"excel-bytes")
            self.assertIn(
                "https://docs.google.com/spreadsheets/d/1s6LB0a_j4k-RhBTKAfZwQVDda25xzgmbGFDTLRC1LvM/export?format=xlsx",
                mocked.call_args.args[0].full_url,
            )


if __name__ == "__main__":
    unittest.main()
