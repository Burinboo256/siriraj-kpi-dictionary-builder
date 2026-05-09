from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.generate_template import SHEET_NAMES, TEMPLATE_HEADERS, create_template_workbook


class TemplateGenerationTests(unittest.TestCase):
    def test_create_template_workbook_has_expected_numbered_sheets_and_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "kpi_dictionary_template.xlsx"

            create_template_workbook(output_path)

            workbook = load_workbook(output_path)
            self.assertEqual(workbook.sheetnames, list(SHEET_NAMES.values()))

            self.assertEqual(
                [cell.value for cell in workbook[SHEET_NAMES["kpi_input_form"]][1]],
                TEMPLATE_HEADERS["KPI_Input_Form"],
            )
            self.assertEqual(
                [cell.value for cell in workbook[SHEET_NAMES["kpi_input_dictionary"]][1]],
                TEMPLATE_HEADERS["KPI_Input_Dictionary"],
            )
            self.assertEqual(
                [cell.value for cell in workbook[SHEET_NAMES["validation_list"]][1]],
                TEMPLATE_HEADERS["Validation_List"],
            )
            self.assertEqual(workbook[SHEET_NAMES["kpi_master"]].sheet_state, "hidden")
            self.assertEqual(workbook[SHEET_NAMES["kpi_logic"]].sheet_state, "hidden")
            self.assertEqual(workbook[SHEET_NAMES["kpi_codeset"]].sheet_state, "hidden")
            self.assertEqual(workbook[SHEET_NAMES["kpi_reference"]].sheet_state, "hidden")
            self.assertEqual(workbook[SHEET_NAMES["kpi_version"]].sheet_state, "hidden")
            self.assertEqual(workbook[SHEET_NAMES["kpi_owner"]].sheet_state, "hidden")

    def test_template_contains_instruction_sheet_examples_required_highlights_and_dropdown_validation(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "kpi_dictionary_template.xlsx"

            create_template_workbook(output_path)

            workbook = load_workbook(output_path)
            instruction_sheet = workbook[SHEET_NAMES["instruction"]]
            input_sheet = workbook[SHEET_NAMES["kpi_input_form"]]
            dictionary_sheet = workbook[SHEET_NAMES["kpi_input_dictionary"]]
            validation_sheet = workbook[SHEET_NAMES["validation_list"]]

            self.assertIn("วิธีกรอก", instruction_sheet["A1"].value)
            self.assertEqual(input_sheet["A2"].value, "Cardiovascular disease (Heart disease: H)")
            self.assertEqual(input_sheet["B2"].value, "Acute coronary syndrome (ACS)")
            self.assertEqual(input_sheet["C2"].value, "DH0101")
            self.assertEqual(input_sheet["I2"].value, "a = จำนวนครั้งของการจำหน่ายด้วยการเสียชีวิตของผู้ป่วย ACS จากทุกหอผู้ป่วยในเดือนนั้น")
            self.assertEqual(input_sheet["J2"].value, "b = จำนวนครั้งของการจำหน่ายทุกสถานะของผู้ป่วย ACS จากทุกหอผู้ป่วยในช่วงเดือนเดียวกันนั้น")
            self.assertEqual(dictionary_sheet["A2"].value, "KPI_Category")
            self.assertEqual(dictionary_sheet["B2"].value, "หมวดตัวชี้วัด")
            self.assertEqual(dictionary_sheet["C2"].value, "Required")
            self.assertIn("ใช้จัดกลุ่มตัวชี้วัด", dictionary_sheet["D2"].value)
            self.assertEqual(dictionary_sheet.max_row, len(TEMPLATE_HEADERS["KPI_Input_Form"]) + 1)
            self.assertEqual(validation_sheet["A2"].value, "KPI_Category")
            self.assertGreater(len(input_sheet.data_validations.dataValidation), 0)

            required_fill = "00FFF2CC"
            optional_fill = input_sheet["M1"].fill.fill_type
            for coordinate in ("A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "K1", "L1"):
                self.assertEqual(input_sheet[coordinate].fill.fgColor.rgb, required_fill)
            self.assertIsNone(optional_fill)


if __name__ == "__main__":
    unittest.main()
