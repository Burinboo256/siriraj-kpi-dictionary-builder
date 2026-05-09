from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

SHEET_NAMES = {
    "instruction": "01_Instruction",
    "kpi_input_form": "02_KPI_Input_Form",
    "kpi_input_dictionary": "03_KPI_Input_Dictionary",
    "kpi_master": "04_KPI_Master",
    "kpi_logic": "05_KPI_Logic",
    "kpi_codeset": "06_KPI_CodeSet",
    "kpi_reference": "07_KPI_Reference",
    "kpi_version": "08_KPI_Version",
    "kpi_owner": "09_KPI_Owner",
    "validation_list": "10_Validation_List",
    "config": "11_Config",
}

REQUIRED_INPUT_HEADERS = [
    "KPI_Category",
    "KPI_Type",
    "KPI_Code",
    "KPI_Name_TH",
    "KPI_Name_EN",
    "Definition_TH",
    "Objective_TH",
    "Formula",
    "Numerator_Definition",
    "Denominator_Definition",
    "Numerator_CodeSet",
    "Denominator_CodeSet",
]

OPTIONAL_INPUT_HEADERS = [
    "Frequency",
    "Benchmark",
    "Interpretation_TH",
    "Reference_Team",
    "Data_Source",
    "Effective_Date",
    "Last_Update_Date",
    "Revision_Reason",
    "Note",
]

TEMPLATE_HEADERS = {
    "Instruction": ["Section", "Detail"],
    "KPI_Input_Form": REQUIRED_INPUT_HEADERS + OPTIONAL_INPUT_HEADERS,
    "KPI_Input_Dictionary": ["Column_Name", "Display_Name_TH", "Required", "Description", "Example", "Allowed_Values"],
    "KPI_Master": [
        "KPI_Category",
        "KPI_Type",
        "KPI_Code",
        "KPI_Short_Name",
        "KPI_Group",
        "Clinical_Area",
        "Service_Area",
        "Direction",
        "Unit",
        "Status",
        "Objective_EN",
        "Definition_EN",
        "Interpretation_EN",
        "Related_KPI_Code",
        "User_Role_Target",
        "Dashboard_URL",
        "Confluence_URL",
    ],
    "KPI_Logic": [
        "KPI_Code",
        "Unit_of_Analysis",
        "Numerator_Logic",
        "Denominator_Logic",
        "Population_Logic",
        "Inclusion_Criteria",
        "Exclusion_Criteria",
        "Data_Source_Table",
        "Join_Key",
        "SQL_Where_Clause",
        "Data_Quality_Check",
        "Known_Limitation",
    ],
    "KPI_CodeSet": [
        "KPI_Code",
        "Used_In",
        "Code_System",
        "Code",
        "Code_Description",
        "Diagnosis_Position",
        "Include_Exclude",
        "Note",
    ],
    "KPI_Reference": [
        "KPI_Code",
        "Reference_Text",
        "Reference_URL",
    ],
    "KPI_Version": [
        "KPI_Code",
        "Version",
        "Changed_By",
        "Approved_By",
    ],
    "KPI_Owner": [
        "KPI_Code",
        "Owner_Department",
        "Owner_Name",
        "Owner_Email",
    ],
    "Validation_List": ["List_Name", "Allowed_Value"],
    "Config": ["Key", "Value"],
}

INPUT_FIELD_METADATA = {
    "KPI_Category": {
        "display_name_th": "หมวดตัวชี้วัด",
        "description": "ใช้จัดกลุ่มตัวชี้วัดในระดับหมวดหลัก เพื่อแสดงหัวข้อในหน้า KPI detail และใช้ทำดัชนีค้นหา",
        "allowed_values_key": "KPI_Category",
    },
    "KPI_Type": {
        "display_name_th": "ประเภทตัวชี้วัด",
        "description": "ใช้ระบุประเภทหรือชุดตัวชี้วัดย่อยภายในหมวดเดียวกัน",
        "allowed_values_key": "KPI_Type",
    },
    "KPI_Code": {
        "display_name_th": "รหัสตัวชี้วัด",
        "description": "รหัสอ้างอิงของ KPI สำหรับใช้งานในเอกสารและเชื่อมข้อมูลข้ามชีต",
    },
    "KPI_Name_TH": {
        "display_name_th": "ชื่อตัวชี้วัด (ภาษาไทย)",
        "description": "ชื่อไทยหลักของตัวชี้วัดที่จะแสดงในหน้า KPI detail",
    },
    "KPI_Name_EN": {
        "display_name_th": "ชื่อตัวชี้วัด (ภาษาอังกฤษ)",
        "description": "ชื่อภาษาอังกฤษของตัวชี้วัดสำหรับแสดงคู่กับชื่อไทย",
    },
    "Definition_TH": {
        "display_name_th": "นิยาม / คำอธิบาย / ความหมาย",
        "description": "อธิบายนิยามของตัวชี้วัด เงื่อนไขสำคัญ และความหมายของการนับข้อมูล",
    },
    "Objective_TH": {
        "display_name_th": "วัตถุประสงค์ของตัวชี้วัด",
        "description": "ระบุว่าตัวชี้วัดนี้ใช้เพื่อประเมินหรือเฝ้าระวังอะไร",
    },
    "Formula": {
        "display_name_th": "สูตรในการคำนวณ",
        "description": "สูตรคำนวณ KPI เช่น (a/b) x 100 หรืออัตราต่อ 1,000 วันนอน",
    },
    "Numerator_Definition": {
        "display_name_th": "ข้อมูลที่ต้องการ: ตัวตั้ง",
        "description": "คำอธิบายข้อมูลตัวตั้งที่ต้องนับจริงใน KPI",
    },
    "Denominator_Definition": {
        "display_name_th": "ข้อมูลที่ต้องการ: ตัวหาร",
        "description": "คำอธิบายข้อมูลตัวหารที่ต้องนับจริงใน KPI",
    },
    "Numerator_CodeSet": {
        "display_name_th": "รหัสโรค/หัตถการที่เกี่ยวข้อง: ตัวตั้ง",
        "description": "ระบุรหัสโรค รหัสหัตถการ หรือเงื่อนไขที่ใช้กับตัวตั้ง",
    },
    "Denominator_CodeSet": {
        "display_name_th": "รหัสโรค/หัตถการที่เกี่ยวข้อง: ตัวหาร",
        "description": "ระบุรหัสโรค รหัสหัตถการ หรือเงื่อนไขที่ใช้กับตัวหาร",
    },
    "Frequency": {
        "display_name_th": "ความถี่ในการจัดทำข้อมูล",
        "description": "ระบุรอบการติดตามหรือความถี่ของตัวชี้วัด",
        "allowed_values_key": "Frequency",
    },
    "Benchmark": {
        "display_name_th": "Benchmark",
        "description": "เกณฑ์เปรียบเทียบหรือเป้าหมายของตัวชี้วัด",
    },
    "Interpretation_TH": {
        "display_name_th": "วิธีการแปลผล",
        "description": "คำอธิบายว่าค่ามากหรือน้อยดีกว่า และเกณฑ์ที่คาดหวัง",
    },
    "Reference_Team": {
        "display_name_th": "ทีม/Reference",
        "description": "หน่วยงานหรือเอกสารอ้างอิงหลักของตัวชี้วัด",
    },
    "Data_Source": {
        "display_name_th": "แหล่งข้อมูล",
        "description": "ระบบหรือแหล่งข้อมูลที่ใช้ดึงข้อมูลมาคำนวณ KPI",
    },
    "Effective_Date": {
        "display_name_th": "วัน เดือน ปี ที่เริ่มใช้",
        "description": "วันที่เริ่มใช้ KPI หรือนิยามชุดนี้",
    },
    "Last_Update_Date": {
        "display_name_th": "วัน เดือน ปี ที่ปรับปรุงครั้งล่าสุด",
        "description": "วันที่มีการปรับปรุง KPI ล่าสุด",
    },
    "Revision_Reason": {
        "display_name_th": "เหตุผลของการปรับปรุง",
        "description": "อธิบายสาเหตุของการแก้ไขนิยามหรือสูตร KPI",
    },
    "Note": {
        "display_name_th": "หมายเหตุ",
        "description": "ข้อมูลเพิ่มเติมที่ควรแสดงในหน้า KPI detail หากมี",
    },
}

DEFAULT_CONFIG = {
    "Document_Title": "KPI Dictionary",
    "Document_Subtitle": "บัญชีตัวชี้วัดเปรียบเทียบ ปีงบประมาณ 2566",
    "Organization_Name": "Faculty of Medicine Siriraj Hospital",
    "Version": "3.0",
    "Generated_By": "SiData+",
    "Font_TH": "TH Sarabun New",
    "Font_EN": "Arial",
    "Page_Size": "A4",
    "Logo_Path": "assets/logo.png",
}

VALIDATION_VALUES = {
    "KPI_Category": [
        "Cardiovascular disease (Heart disease: H)",
        "Infection prevention",
        "Medication safety",
    ],
    "KPI_Type": [
        "Acute coronary syndrome (ACS)",
        "Surgical antibiotic prophylaxis",
        "Medication reconciliation",
    ],
    "Frequency": ["ทุกเดือน", "ทุก 3 เดือน", "ทุก 6 เดือน", "ทุกปี"],
    "KPI_Group": ["Disease", "Process", "Safety", "Corporate"],
    "Service_Area": ["IPD", "OPD", "ER", "OR", "Hospital-wide"],
    "Direction": ["Higher is better", "Lower is better", "Target range"],
    "Owner_Department": ["Cardiology", "Infection Control", "Pharmacy"],
}

INSTRUCTION_ROWS = [
    ("วิธีกรอก", "กรอกเฉพาะชีต 02_KPI_Input_Form เป็นหลัก โดยช่องสีเหลืองคือ required และช่องที่ไม่ไฮไลต์คือ optional"),
    ("โครงสร้างข้อมูล", "แบบฟอร์มหลักออกแบบตามหน้า KPI detail reference เพื่อให้ผู้ใช้งานกรอกเฉพาะข้อมูลที่ใช้แสดงผลจริง"),
    ("dictionary ของคอลัมน์", "ดูความหมาย วิธีกรอก และตัวอย่างของแต่ละคอลัมน์ได้ในชีต 03_KPI_Input_Dictionary"),
    ("ชีตเสริม", "ชีต 04-09 เป็นข้อมูลสำหรับ analyst/admin เท่านั้น และถูกซ่อนไว้เพื่อไม่ให้รบกวนผู้ใช้งานทั่วไป"),
    ("รหัสโรค/หัตถการ", "กรอกเป็นข้อความรวมได้ในคอลัมน์ Numerator_CodeSet และ Denominator_CodeSet หากยังไม่ต้องการแตกเป็นหลายบรรทัด"),
    ("วันที่", "กรอกวันที่ในรูปแบบที่หน่วยงานใช้งานจริงได้ เช่น 1 ตุลาคม 2567 หรือ 2024-10-01"),
]

INPUT_FORM_EXAMPLES = [
    {
        "KPI_Category": "Cardiovascular disease (Heart disease: H)",
        "KPI_Type": "Acute coronary syndrome (ACS)",
        "KPI_Code": "DH0101",
        "KPI_Name_TH": "ร้อยละการเสียชีวิตของผู้ป่วยภาวะหัวใจขาดเลือดเฉียบพลัน",
        "KPI_Name_EN": "Acute coronary syndrome: Percent of mortality",
        "Definition_TH": "1. ผู้ป่วย Acute coronary syndrome (ACS) หมายถึง ผู้ป่วยใน อายุ ≥ 18 ปี ที่มี Principal diagnosis (Pdx) เป็นภาวะหัวใจขาดเลือดเฉียบพลัน ได้แก่ STEMI หรือ NSTE-ACS 2. การเสียชีวิตของผู้ป่วย ACS หมายถึง การเสียชีวิตจากทุกสาเหตุ",
        "Objective_TH": "ประเมินประสิทธิผลกระบวนการดูแลรักษากลุ่มผู้ป่วย ACS",
        "Formula": "(a/b) x 100",
        "Numerator_Definition": "a = จำนวนครั้งของการจำหน่ายด้วยการเสียชีวิตของผู้ป่วย ACS จากทุกหอผู้ป่วยในเดือนนั้น",
        "Denominator_Definition": "b = จำนวนครั้งของการจำหน่ายทุกสถานะของผู้ป่วย ACS จากทุกหอผู้ป่วยในช่วงเดือนเดียวกันนั้น",
        "Numerator_CodeSet": "Pdx = I21.0, I21.1, I21.2, I21.3, I21.4, I21.9 ที่เสียชีวิตจากทุกสาเหตุ",
        "Denominator_CodeSet": "Pdx = I21.0, I21.1, I21.2, I21.3, I21.4, I21.9",
        "Frequency": "ทุกเดือน",
        "Benchmark": "≤ 9",
        "Interpretation_TH": "ค่ายิ่งน้อย = มีคุณภาพดี เป้าหมาย ≤ 9",
        "Reference_Team": "NQF, THIP I, แนวทางเวชปฏิบัติการดูแลรักษาผู้ป่วยภาวะหัวใจขาดเลือดเฉียบพลัน พ.ศ. 2563",
        "Data_Source": "ข้อมูลจำหน่ายผู้ป่วยใน / ICD-10 coding",
        "Effective_Date": "1 ตุลาคม 2557",
        "Last_Update_Date": "1 ตุลาคม 2564",
        "Revision_Reason": "ปรับนิยามสอดคล้องตามแนวทางเวชปฏิบัติปี 2563",
        "Note": "",
    },
    {
        "KPI_Category": "Infection prevention",
        "KPI_Type": "Surgical antibiotic prophylaxis",
        "KPI_Code": "IP0201",
        "KPI_Name_TH": "ร้อยละผู้ป่วยได้รับยาปฏิชีวนะก่อนผ่าตัดภายในเวลาที่กำหนด",
        "KPI_Name_EN": "Percent of patients receiving prophylactic antibiotic on time",
        "Definition_TH": "ผู้ป่วยที่ได้รับการผ่าตัดชนิดที่ต้องได้รับ prophylactic antibiotic และได้รับยาก่อนลงมีดในช่วงเวลาที่กำหนด",
        "Objective_TH": "เพื่อให้ผู้ป่วยได้รับยาปฏิชีวนะก่อนผ่าตัดตรงเวลา",
        "Formula": "(a/b) x 100",
        "Numerator_Definition": "a = จำนวนผ่าตัดที่ได้รับ prophylactic antibiotic ภายในเวลาที่กำหนดก่อน incision",
        "Denominator_Definition": "b = จำนวนผ่าตัดที่เข้าเกณฑ์ต้องได้รับ prophylactic antibiotic ทั้งหมด",
        "Numerator_CodeSet": "Procedure กลุ่ม clean-contaminated surgery ที่ได้รับยาใน 60 นาทีก่อน incision",
        "Denominator_CodeSet": "Procedure กลุ่ม clean-contaminated surgery ทั้งหมด",
        "Frequency": "ทุกเดือน",
        "Benchmark": "≥ 95%",
        "Interpretation_TH": "ค่ายิ่งมาก = มีคุณภาพดี เป้าหมาย ≥ 95%",
        "Reference_Team": "WHO surgical safety guidance / Infection prevention team",
        "Data_Source": "OR checklist / medication administration record",
        "Effective_Date": "1 ตุลาคม 2567",
        "Last_Update_Date": "1 ตุลาคม 2567",
        "Revision_Reason": "เริ่มใช้งานตัวชี้วัด",
        "Note": "",
    },
    {
        "KPI_Category": "Medication safety",
        "KPI_Type": "Medication reconciliation",
        "KPI_Code": "PS0301",
        "KPI_Name_TH": "ร้อยละความครบถ้วนของการทบทวนรายการยาก่อนจำหน่าย",
        "KPI_Name_EN": "Medication reconciliation completion rate before discharge",
        "Definition_TH": "การทบทวนรายการยาก่อนจำหน่ายครบถ้วนสำหรับผู้ป่วยที่มียากลับบ้าน",
        "Objective_TH": "เพื่อลดความคลาดเคลื่อนทางยาก่อนจำหน่าย",
        "Formula": "(a/b) x 100",
        "Numerator_Definition": "a = จำนวนจำหน่ายที่มีการทบทวนรายการยาครบถ้วน",
        "Denominator_Definition": "b = จำนวนจำหน่ายที่มียากลับบ้านทั้งหมด",
        "Numerator_CodeSet": "MEDREC-DC completed",
        "Denominator_CodeSet": "Discharge with take-home medication",
        "Frequency": "ทุกเดือน",
        "Benchmark": "≥ 90%",
        "Interpretation_TH": "ค่ายิ่งมาก = มีคุณภาพดี เป้าหมาย ≥ 90%",
        "Reference_Team": "Hospital medication reconciliation standard / Pharmacy team",
        "Data_Source": "EMR medication reconciliation module",
        "Effective_Date": "1 ตุลาคม 2567",
        "Last_Update_Date": "1 มีนาคม 2568",
        "Revision_Reason": "เพิ่มนิยามการนับ discharge",
        "Note": "",
    },
]

MASTER_EXAMPLES = [
    {
        "KPI_Category": "Cardiovascular disease (Heart disease: H)",
        "KPI_Type": "Acute coronary syndrome (ACS)",
        "KPI_Code": "DH0101",
        "KPI_Short_Name": "ACS mortality",
        "KPI_Group": "Disease",
        "Clinical_Area": "Cardiology",
        "Service_Area": "IPD",
        "Direction": "Lower is better",
        "Unit": "Percent",
        "Status": "Active",
        "Objective_EN": "To monitor ACS mortality outcome.",
        "Definition_EN": "Mortality rate of ACS discharges.",
        "Interpretation_EN": "Lower is better. Target ≤ 9.",
        "Related_KPI_Code": "",
        "User_Role_Target": "Executive, Cardiologist",
        "Dashboard_URL": "",
        "Confluence_URL": "",
    },
    {
        "KPI_Category": "Infection prevention",
        "KPI_Type": "Surgical antibiotic prophylaxis",
        "KPI_Code": "IP0201",
        "KPI_Short_Name": "On-time prophylactic antibiotic",
        "KPI_Group": "Process",
        "Clinical_Area": "Surgery",
        "Service_Area": "OR",
        "Direction": "Higher is better",
        "Unit": "Percent",
        "Status": "Active",
        "Objective_EN": "To ensure timely prophylactic antibiotic administration.",
        "Definition_EN": "Percent of eligible patients receiving prophylactic antibiotics on time.",
        "Interpretation_EN": "Higher is better. Target ≥ 95%.",
        "Related_KPI_Code": "",
        "User_Role_Target": "Surgeon, OR nurse",
        "Dashboard_URL": "",
        "Confluence_URL": "",
    },
    {
        "KPI_Category": "Medication safety",
        "KPI_Type": "Medication reconciliation",
        "KPI_Code": "PS0301",
        "KPI_Short_Name": "Medication reconciliation",
        "KPI_Group": "Safety",
        "Clinical_Area": "Pharmacy",
        "Service_Area": "Hospital-wide",
        "Direction": "Higher is better",
        "Unit": "Percent",
        "Status": "Active",
        "Objective_EN": "To reduce medication discrepancies before discharge.",
        "Definition_EN": "Percent of completed medication reconciliation before discharge.",
        "Interpretation_EN": "Higher is better. Target ≥ 90%.",
        "Related_KPI_Code": "",
        "User_Role_Target": "Pharmacist, Ward nurse",
        "Dashboard_URL": "",
        "Confluence_URL": "",
    },
]

LOGIC_EXAMPLES = [
    {
        "KPI_Code": "DH0101",
        "Unit_of_Analysis": "Admission",
        "Numerator_Logic": "Pdx in I21.* and discharge status = death",
        "Denominator_Logic": "Pdx in I21.*",
        "Population_Logic": "Discharged inpatient with ACS diagnosis",
        "Inclusion_Criteria": "ผู้ป่วยในอายุ ≥ 18 ปี ที่มีการวินิจฉัย ACS",
        "Exclusion_Criteria": "ส่งต่อออกก่อนสิ้นสุดการรักษา",
        "Data_Source_Table": "fact_inpatient_discharge",
        "Join_Key": "admission_id",
        "SQL_Where_Clause": "principal_dx like 'I21%'",
        "Data_Quality_Check": "Check discharge status completeness",
        "Known_Limitation": "Depends on diagnosis coding quality",
    },
    {
        "KPI_Code": "IP0201",
        "Unit_of_Analysis": "Operation",
        "Numerator_Logic": "antibiotic_time within 60 mins before incision",
        "Denominator_Logic": "eligible_prophylaxis_flag = 1",
        "Population_Logic": "Elective surgery eligible for prophylactic antibiotic",
        "Inclusion_Criteria": "Elective clean-contaminated surgery",
        "Exclusion_Criteria": "Allergy with alternative timing",
        "Data_Source_Table": "fact_or_case",
        "Join_Key": "operation_id",
        "SQL_Where_Clause": "eligible_prophylaxis_flag = 1",
        "Data_Quality_Check": "Check incision time and medication time completeness",
        "Known_Limitation": "Manual checklist may lag",
    },
    {
        "KPI_Code": "PS0301",
        "Unit_of_Analysis": "Discharge",
        "Numerator_Logic": "med_rec_completed_flag = 1",
        "Denominator_Logic": "discharge_med_count > 0",
        "Population_Logic": "Inpatient discharge with prescribed medications",
        "Inclusion_Criteria": "ผู้ป่วยจำหน่ายที่มียากลับบ้าน",
        "Exclusion_Criteria": "เสียชีวิตในโรงพยาบาล",
        "Data_Source_Table": "fact_medication_reconciliation",
        "Join_Key": "visit_id",
        "SQL_Where_Clause": "discharge_med_count > 0",
        "Data_Quality_Check": "Check completion status and pharmacist assignment",
        "Known_Limitation": "Dependent on EMR workflow adoption",
    },
]

CODESET_EXAMPLES = [
    ("DH0101", "Numerator", "ICD-10", "I21.0-I21.9", "ACS discharge with death", "Pdx", "Include", ""),
    ("DH0101", "Denominator", "ICD-10", "I21.0-I21.9", "ACS discharge all status", "Pdx", "Include", ""),
    ("IP0201", "Numerator", "Procedure", "OR-ABX-01", "Antibiotic within 60 minutes before incision", "Any", "Include", ""),
    ("IP0201", "Denominator", "Procedure", "OR-ELIGIBLE-01", "Eligible surgery requiring prophylactic antibiotic", "Any", "Include", ""),
    ("PS0301", "Numerator", "Procedure", "MEDREC-DC", "Medication reconciliation before discharge", "Any", "Include", ""),
    ("PS0301", "Denominator", "Encounter", "DC-MED", "Discharge with take-home medication", "Any", "Include", ""),
]

REFERENCE_EXAMPLES = [
    ("DH0101", "NQF, THIP I, แนวทางเวชปฏิบัติการดูแลรักษาผู้ป่วยภาวะหัวใจขาดเลือดเฉียบพลัน พ.ศ. 2563", ""),
    ("IP0201", "WHO surgical safety and prophylactic antibiotic guidance", ""),
    ("PS0301", "Hospital medication reconciliation standard", ""),
]

VERSION_EXAMPLES = [
    ("DH0101", "1.2", "Data Analyst A", "Medical Director"),
    ("IP0201", "1.0", "QI Team", "Chief of Surgery"),
    ("PS0301", "1.1", "Pharmacy Analyst", "Chief Pharmacist"),
]

OWNER_EXAMPLES = [
    ("DH0101", "Cardiology", "Dr. Somchai", "somchai@example.org"),
    ("IP0201", "Infection Control", "Ms. Suda", "suda@example.org"),
    ("PS0301", "Pharmacy", "Mr. Anan", "anan@example.org"),
]


def _build_input_dictionary_rows() -> list[dict[str, str]]:
    sample_row = INPUT_FORM_EXAMPLES[0]
    rows: list[dict[str, str]] = []
    for header in TEMPLATE_HEADERS["KPI_Input_Form"]:
        metadata = INPUT_FIELD_METADATA[header]
        allowed_key = metadata.get("allowed_values_key", "")
        allowed_values = ", ".join(VALIDATION_VALUES.get(allowed_key, [])) if allowed_key else ""
        rows.append(
            {
                "Column_Name": header,
                "Display_Name_TH": metadata["display_name_th"],
                "Required": "Required" if header in REQUIRED_INPUT_HEADERS else "Optional",
                "Description": metadata["description"],
                "Example": sample_row.get(header, ""),
                "Allowed_Values": allowed_values,
            }
        )
    return rows


INPUT_DICTIONARY_ROWS = _build_input_dictionary_rows()

ADVANCED_SHEETS = {
    SHEET_NAMES["kpi_master"],
    SHEET_NAMES["kpi_logic"],
    SHEET_NAMES["kpi_codeset"],
    SHEET_NAMES["kpi_reference"],
    SHEET_NAMES["kpi_version"],
    SHEET_NAMES["kpi_owner"],
}


def _style_header_row(worksheet, required_headers: set[str] | None = None) -> None:
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    admin_fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in worksheet[1]:
        cell.font = font
        cell.alignment = alignment
        if required_headers is not None:
            if cell.value in required_headers:
                cell.fill = required_fill
        else:
            cell.fill = admin_fill


def _autosize_columns(worksheet) -> None:
    for index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 18), 42)


def _append_dict_rows(worksheet, headers: list[str], rows: list[dict[str, str]]) -> None:
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])


def _append_tuple_rows(worksheet, rows: list[tuple[str, ...]]) -> None:
    for row in rows:
        worksheet.append(list(row))


def _add_validation_lists(workbook: Workbook) -> None:
    sheet = workbook[SHEET_NAMES["validation_list"]]
    current_row = 2
    for list_name, values in VALIDATION_VALUES.items():
        start_row = current_row
        for value in values:
            sheet.append([list_name, value])
            current_row += 1
        end_row = current_row - 1
        defined_name = DefinedName(
            list_name,
            attr_text=f"'{SHEET_NAMES['validation_list']}'!$B${start_row}:$B${end_row}",
        )
        workbook.defined_names.add(defined_name)


def _add_dropdown_validations(workbook: Workbook) -> None:
    validation_targets = {
        "KPI_Category": SHEET_NAMES["kpi_input_form"],
        "KPI_Type": SHEET_NAMES["kpi_input_form"],
        "Frequency": SHEET_NAMES["kpi_input_form"],
        "KPI_Group": SHEET_NAMES["kpi_master"],
        "Service_Area": SHEET_NAMES["kpi_master"],
        "Direction": SHEET_NAMES["kpi_master"],
        "Owner_Department": SHEET_NAMES["kpi_owner"],
    }
    for field, sheet_name in validation_targets.items():
        headers = list(workbook[sheet_name][1])
        column_index = next((index for index, cell in enumerate(headers, start=1) if cell.value == field), None)
        if column_index is None:
            continue
        column_letter = get_column_letter(column_index)
        validation = DataValidation(type="list", formula1=f"={field}", allow_blank=True, showDropDown=True)
        validation.prompt = f"เลือกค่าจากรายการ {field}"
        validation.error = f"ค่า {field} ไม่อยู่ในรายการที่กำหนด"
        validation.add(f"{column_letter}2:{column_letter}500")
        workbook[sheet_name].add_data_validation(validation)


def create_template_workbook(output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    instruction_sheet = workbook.active
    instruction_sheet.title = SHEET_NAMES["instruction"]

    for sheet_name in SHEET_NAMES.values():
        if sheet_name == SHEET_NAMES["instruction"]:
            continue
        workbook.create_sheet(sheet_name)

    sheet_header_map = {
        SHEET_NAMES["kpi_input_form"]: TEMPLATE_HEADERS["KPI_Input_Form"],
        SHEET_NAMES["kpi_input_dictionary"]: TEMPLATE_HEADERS["KPI_Input_Dictionary"],
        SHEET_NAMES["kpi_master"]: TEMPLATE_HEADERS["KPI_Master"],
        SHEET_NAMES["kpi_logic"]: TEMPLATE_HEADERS["KPI_Logic"],
        SHEET_NAMES["kpi_codeset"]: TEMPLATE_HEADERS["KPI_CodeSet"],
        SHEET_NAMES["kpi_reference"]: TEMPLATE_HEADERS["KPI_Reference"],
        SHEET_NAMES["kpi_version"]: TEMPLATE_HEADERS["KPI_Version"],
        SHEET_NAMES["kpi_owner"]: TEMPLATE_HEADERS["KPI_Owner"],
        SHEET_NAMES["validation_list"]: TEMPLATE_HEADERS["Validation_List"],
        SHEET_NAMES["config"]: TEMPLATE_HEADERS["Config"],
    }

    for sheet_name, headers in sheet_header_map.items():
        sheet = workbook[sheet_name]
        sheet.append(headers)
        if sheet_name == SHEET_NAMES["kpi_input_form"]:
            _style_header_row(sheet, required_headers=set(REQUIRED_INPUT_HEADERS))
        else:
            _style_header_row(sheet)
        sheet.freeze_panes = "A2"

    _append_tuple_rows(instruction_sheet, INSTRUCTION_ROWS)
    _append_dict_rows(workbook[SHEET_NAMES["kpi_input_form"]], TEMPLATE_HEADERS["KPI_Input_Form"], INPUT_FORM_EXAMPLES)
    _append_dict_rows(workbook[SHEET_NAMES["kpi_input_dictionary"]], TEMPLATE_HEADERS["KPI_Input_Dictionary"], INPUT_DICTIONARY_ROWS)
    _append_dict_rows(workbook[SHEET_NAMES["kpi_master"]], TEMPLATE_HEADERS["KPI_Master"], MASTER_EXAMPLES)
    _append_dict_rows(workbook[SHEET_NAMES["kpi_logic"]], TEMPLATE_HEADERS["KPI_Logic"], LOGIC_EXAMPLES)
    _append_tuple_rows(workbook[SHEET_NAMES["kpi_codeset"]], CODESET_EXAMPLES)
    _append_tuple_rows(workbook[SHEET_NAMES["kpi_reference"]], REFERENCE_EXAMPLES)
    _append_tuple_rows(workbook[SHEET_NAMES["kpi_version"]], VERSION_EXAMPLES)
    _append_tuple_rows(workbook[SHEET_NAMES["kpi_owner"]], OWNER_EXAMPLES)
    _add_validation_lists(workbook)

    config_sheet = workbook[SHEET_NAMES["config"]]
    for key, value in DEFAULT_CONFIG.items():
        config_sheet.append([key, value])

    _add_dropdown_validations(workbook)

    for sheet_name in ADVANCED_SHEETS:
        workbook[sheet_name].sheet_state = "hidden"

    for sheet_name in workbook.sheetnames:
        _autosize_columns(workbook[sheet_name])

    workbook.save(output_path)
    return output_path


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parents[1]
    create_template_workbook(project_root / "input" / "kpi_dictionary_template.xlsx")
