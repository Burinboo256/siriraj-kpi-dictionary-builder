from __future__ import annotations

from collections import Counter, defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
import re
import ssl
from typing import Any
from urllib.error import URLError
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook

from src.generate_template import DEFAULT_CONFIG, SHEET_NAMES, VALIDATION_VALUES

INDEX_DEFINITIONS = {
    "Index by KPI Code": ("KPI_Code",),
    "Index by KPI Category": ("KPI_Category",),
    "Index by KPI Type": ("KPI_Type",),
    "Index by Frequency": ("Frequency",),
    "Index by Reference Team": ("Reference_Team",),
    "Index by ICD / Procedure Code": ("ICD_Procedure_Index",),
}

REQUIRED_FIELDS = {
    "KPI_Code": "Missing required fields",
    "KPI_Name_TH": "Missing required fields",
    "KPI_Category": "Missing required fields",
    "KPI_Type": "Missing required fields",
    "Definition_TH": "Missing required fields",
    "Formula": "Missing required fields",
    "Numerator_Definition": "Missing required fields",
    "Denominator_Definition": "Missing required fields",
}

GOOGLE_SHEET_ID_PATTERN = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
CRITICAL_ISSUE_TYPES = {"Missing required fields", "Missing formula"}
VALIDATION_FIELD_LABELS = {
    "KPI_Category": "หมวดตัวชี้วัด",
    "KPI_Type": "ประเภทตัวชี้วัด",
    "KPI_Code": "รหัสตัวชี้วัด",
    "KPI_Name_TH": "ชื่อตัวชี้วัด (ภาษาไทย)",
    "KPI_Name_EN": "ชื่อตัวชี้วัด (ภาษาอังกฤษ)",
    "Definition_TH": "นิยาม คำอธิบาย ความหมายของตัวชี้วัด",
    "Objective_TH": "วัตถุประสงค์ของตัวชี้วัด",
    "Formula": "สูตรในการคำนวณ",
    "Numerator_Definition": "ข้อมูลที่ต้องการ: ตัวตั้ง",
    "Denominator_Definition": "ข้อมูลที่ต้องการ: ตัวหาร",
    "Numerator_CodeSet": "รหัสโรค/หัตถการที่เกี่ยวข้อง: ตัวตั้ง",
    "Denominator_CodeSet": "รหัสโรค/หัตถการที่เกี่ยวข้อง: ตัวหาร",
}


@dataclass
class KPIDataset:
    records: list[dict[str, str]]
    config: dict[str, str]
    validation_lists: dict[str, list[str]]
    master_code_counts: dict[str, int]


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()
    return str(value).strip()


def _extract_google_sheet_id(source: str) -> str | None:
    if source.startswith("http://") or source.startswith("https://"):
        match = GOOGLE_SHEET_ID_PATTERN.search(source)
        return match.group(1) if match else None
    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", source):
        return source
    return None


def download_google_sheet_workbook(source: str, target_dir: Path) -> Path:
    sheet_id = _extract_google_sheet_id(source)
    if not sheet_id:
        raise ValueError("Invalid Google Sheet URL or sheet ID")

    target_dir.mkdir(parents=True, exist_ok=True)
    output_path = target_dir / f"{sheet_id}.xlsx"
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    request = Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(request) as response:
            output_path.write_bytes(response.read())
    except URLError as error:
        if isinstance(getattr(error, "reason", None), ssl.SSLCertVerificationError):
            insecure_context = ssl._create_unverified_context()
            with urlopen(request, context=insecure_context) as response:
                output_path.write_bytes(response.read())
        else:
            raise
    return output_path


def resolve_workbook_source(source: str | Path, temp_dir: Path) -> Path:
    source_path = Path(source)
    if source_path.exists():
        return source_path

    sheet_id = _extract_google_sheet_id(str(source))
    if sheet_id:
        return download_google_sheet_workbook(str(source), temp_dir)

    raise FileNotFoundError(f"Workbook source not found: {source}")


def _load_with_pandas(workbook_path: Path, sheet_name: str) -> list[dict[str, str]] | None:
    try:
        import pandas as pd
    except Exception:
        return None

    dataframe = pd.read_excel(workbook_path, sheet_name=sheet_name, dtype=str).fillna("")
    return [
        {column: _normalize_value(value) for column, value in record.items()}
        for record in dataframe.to_dict(orient="records")
        if any(_normalize_value(value) for value in record.values())
    ]


def _load_with_openpyxl(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [_normalize_value(value) for value in rows[0]]
    records = []
    for row in rows[1:]:
        record = {header: _normalize_value(value) for header, value in zip(headers, row, strict=False)}
        if any(record.values()):
            records.append(record)
    return records


def _list_sheet_names(workbook_path: Path) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    return list(workbook.sheetnames)


def _infer_direction(text: str) -> str:
    if "ยิ่งน้อย" in text or "lower" in text.lower():
        return "Lower is better"
    if "ยิ่งมาก" in text or "higher" in text.lower():
        return "Higher is better"
    return ""


def _load_legacy_google_sheet_dataset(workbook_path: Path) -> KPIDataset:
    rows = _load_with_openpyxl(workbook_path, "KPIs")
    records: list[dict[str, str]] = []
    for row in rows:
        record = {
            "KPI_Code": row.get("kpi_code", ""),
            "KPI_Name_TH": row.get("name_th", ""),
            "KPI_Name_EN": row.get("name_en", ""),
            "KPI_Category": row.get("category_name", ""),
            "KPI_Type": row.get("type_name", ""),
            "Frequency": row.get("frequency", ""),
            "Direction": _infer_direction(row.get("interpretation", "")),
            "Reference_Team": row.get("reference", ""),
            "Data_Source": "",
            "Effective_Date": row.get("start_date", ""),
            "Objective_TH": row.get("objective", ""),
            "Objective_EN": "",
            "Definition_TH": row.get("definition", ""),
            "Definition_EN": "",
            "Interpretation_TH": row.get("interpretation", ""),
            "Interpretation_EN": "",
            "Benchmark": row.get("benchmark", ""),
            "Note": row.get("notes", ""),
            "Formula": row.get("formula", ""),
            "Numerator_Definition": row.get("numerator_data", ""),
            "Numerator_Logic": row.get("numerator_data", ""),
            "Denominator_Definition": row.get("denominator_data", ""),
            "Denominator_Logic": row.get("denominator_data", ""),
            "Reference_Text_All": row.get("reference", ""),
            "Version": "1.0",
            "Last_Update_Date": row.get("last_update_date", ""),
            "Revision_Reason": row.get("update_reason", ""),
            "Numerator_CodeSet": row.get("numerator_icd", ""),
            "Denominator_CodeSet": row.get("denominator_icd", ""),
            "ICD_Procedure_Index": ", ".join(
                value
                for value in (
                    row.get("numerator_icd", ""),
                    row.get("denominator_icd", ""),
                )
                if value
            ),
        }
        if record["KPI_Code"]:
            records.append(record)

    validation_lists = {key: list(values) for key, values in VALIDATION_VALUES.items()}
    return KPIDataset(
        records=records,
        config=dict(DEFAULT_CONFIG),
        validation_lists=validation_lists,
        master_code_counts=dict(Counter(record["KPI_Code"] for record in records)),
    )


def load_sheet_records(workbook_path: Path, semantic_name: str) -> list[dict[str, str]]:
    workbook_path = Path(workbook_path)
    sheet_name = SHEET_NAMES.get(semantic_name, semantic_name)
    records = _load_with_pandas(workbook_path, sheet_name)
    if records is not None:
        return records
    return _load_with_openpyxl(workbook_path, sheet_name)


def load_config(workbook_path: Path) -> dict[str, str]:
    if SHEET_NAMES["config"] not in _list_sheet_names(workbook_path):
        return dict(DEFAULT_CONFIG)
    rows = load_sheet_records(workbook_path, "config")
    return {row["Key"]: row["Value"] for row in rows if row.get("Key")}


def _load_validation_lists(workbook_path: Path) -> dict[str, list[str]]:
    if SHEET_NAMES["validation_list"] not in _list_sheet_names(workbook_path):
        return {key: list(values) for key, values in VALIDATION_VALUES.items()}
    rows = load_sheet_records(workbook_path, "validation_list")
    grouped: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        if row.get("List_Name") and row.get("Allowed_Value"):
            grouped[row["List_Name"]].append(row["Allowed_Value"])
    return dict(grouped)


def _index_by_code(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {row["KPI_Code"]: dict(row) for row in rows if row.get("KPI_Code")}


def _merge_kpi_records(workbook_path: Path) -> list[dict[str, str]]:
    input_rows = _index_by_code(load_sheet_records(workbook_path, "kpi_input_form"))
    master_rows = _index_by_code(load_sheet_records(workbook_path, "kpi_master"))
    logic_rows = _index_by_code(load_sheet_records(workbook_path, "kpi_logic"))
    version_rows = _index_by_code(load_sheet_records(workbook_path, "kpi_version"))
    owner_rows = _index_by_code(load_sheet_records(workbook_path, "kpi_owner"))

    reference_rows = defaultdict(list)
    for row in load_sheet_records(workbook_path, "kpi_reference"):
        if row.get("KPI_Code"):
            reference_rows[row["KPI_Code"]].append(row)

    code_rows = defaultdict(list)
    for row in load_sheet_records(workbook_path, "kpi_codeset"):
        if row.get("KPI_Code"):
            code_rows[row["KPI_Code"]].append(row)

    all_codes = sorted(input_rows)
    records: list[dict[str, str]] = []
    for kpi_code in all_codes:
        record: dict[str, str] = {"KPI_Code": kpi_code}
        for source in (input_rows.get(kpi_code, {}), master_rows.get(kpi_code, {}), logic_rows.get(kpi_code, {}), version_rows.get(kpi_code, {}), owner_rows.get(kpi_code, {})):
            for key, value in source.items():
                if value:
                    record[key] = value

        numerator_codes = [row.get("Code", "") for row in code_rows.get(kpi_code, []) if row.get("Used_In") == "Numerator" and row.get("Code")]
        denominator_codes = [row.get("Code", "") for row in code_rows.get(kpi_code, []) if row.get("Used_In") == "Denominator" and row.get("Code")]
        all_codes_for_index = [row.get("Code", "") for row in code_rows.get(kpi_code, []) if row.get("Code")]
        if numerator_codes and not record.get("Numerator_CodeSet"):
            record["Numerator_CodeSet"] = ", ".join(numerator_codes)
        if denominator_codes and not record.get("Denominator_CodeSet"):
            record["Denominator_CodeSet"] = ", ".join(denominator_codes)
        record["ICD_Procedure_Index"] = ", ".join(all_codes_for_index)
        record["Reference_Text_All"] = "\n".join(
            value
            for value in (
                record.get("Reference_Team", ""),
                *(row.get("Reference_Text", "") for row in reference_rows.get(kpi_code, [])),
            )
            if value
        )
        record["CodeSet_Rows"] = str(len(code_rows.get(kpi_code, [])))
        records.append(record)
    return records


def load_kpi_dataset(workbook_path: Path) -> KPIDataset:
    sheet_names = set(_list_sheet_names(workbook_path))
    if SHEET_NAMES["kpi_master"] not in sheet_names and "KPIs" in sheet_names:
        return _load_legacy_google_sheet_dataset(workbook_path)

    input_rows = load_sheet_records(workbook_path, "kpi_input_form")
    master_rows = load_sheet_records(workbook_path, "kpi_master")
    code_counts = Counter(row.get("KPI_Code", "") for row in input_rows if row.get("KPI_Code"))
    return KPIDataset(
        records=_merge_kpi_records(workbook_path),
        config=load_config(workbook_path),
        validation_lists=_load_validation_lists(workbook_path),
        master_code_counts=dict(code_counts),
    )


def estimate_detail_page_numbers(records: Iterable[dict[str, str]], starting_page: int = 12) -> list[dict[str, str]]:
    indexed = []
    for offset, record in enumerate(records):
        row = dict(record)
        row["Page_No"] = str(starting_page + offset)
        indexed.append(row)
    return indexed


def build_index_sections(records: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    indexed_records = estimate_detail_page_numbers(records)
    sections: dict[str, list[dict[str, str]]] = {}
    for title, fields in INDEX_DEFINITIONS.items():
        rows: list[dict[str, str]] = []
        for record in indexed_records:
            for field in fields:
                value = record.get(field, "")
                if value:
                    row = dict(record)
                    row["Index_Value"] = value
                    rows.append(row)
        sections[title] = sorted(rows, key=lambda row: (row["Index_Value"], row["KPI_Code"]))
    return sections


def _issue(issue_type: str, kpi_code: str, field_name: str, message: str, sheet_name: str = "") -> dict[str, str]:
    return {
        "Issue_Type": issue_type,
        "KPI_Code": kpi_code,
        "Field_Name": field_name,
        "Message": message,
        "Sheet_Name": sheet_name,
    }


def build_validation_issues(dataset: KPIDataset) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    code_counts = Counter(dataset.master_code_counts)

    for code, count in code_counts.items():
        if count > 1:
            issues.append(_issue("Duplicate KPI_Code", code, "KPI_Code", f"Found {count} records with the same KPI code", SHEET_NAMES["kpi_master"]))

    allowed_values = {key: set(values) for key, values in dataset.validation_lists.items()}
    known_codes = {record.get("KPI_Code", "") for record in dataset.records}

    for record in dataset.records:
        kpi_code = record.get("KPI_Code", "")
        for field_name, issue_type in REQUIRED_FIELDS.items():
            if not record.get(field_name):
                issues.append(_issue(issue_type, kpi_code, field_name, f"{field_name} is required", SHEET_NAMES["kpi_master"]))

        for field_name in ("KPI_Category", "KPI_Type", "Frequency", "KPI_Group", "Service_Area", "Direction", "Owner_Department"):
            value = record.get(field_name, "")
            if value and value not in allowed_values.get(field_name, set(VALIDATION_VALUES.get(field_name, []))):
                issues.append(_issue("Invalid dropdown values", kpi_code, field_name, f"{value} is not in the allowed list", SHEET_NAMES["kpi_input_form"]))

        if not record.get("Formula"):
            issues.append(_issue("Missing formula", kpi_code, "Formula", "Formula is blank", SHEET_NAMES["kpi_input_form"]))
        if record.get("ICD_Procedure_Index") and not any(char.isdigit() for char in record["ICD_Procedure_Index"]):
            issues.append(_issue("ICD format warning", kpi_code, "ICD_Procedure_Index", "ICD / procedure codes look malformed", SHEET_NAMES["kpi_codeset"]))

        related_code = record.get("Related_KPI_Code", "")
        if related_code and related_code not in known_codes:
            issues.append(_issue("Related KPI code not found in KPI_Master", kpi_code, "Related_KPI_Code", f"{related_code} was not found", SHEET_NAMES["kpi_master"]))

    return issues


def build_kpi_validation_summaries(
    records: list[dict[str, str]],
    issues: list[dict[str, str]],
) -> dict[str, dict[str, str | list[str]]]:
    issues_by_code: dict[str, list[dict[str, str]]] = defaultdict(list)
    for issue in issues:
        kpi_code = issue.get("KPI_Code", "")
        if kpi_code:
            issues_by_code[kpi_code].append(issue)

    summaries: dict[str, dict[str, str | list[str]]] = {}
    for record in records:
        kpi_code = record.get("KPI_Code", "")
        record_issues = issues_by_code.get(kpi_code, [])
        critical = [issue for issue in record_issues if issue.get("Issue_Type") in CRITICAL_ISSUE_TYPES]

        if critical:
            status = "Incomplete"
        else:
            status = "Complete"

        summary_parts: list[str] = []
        if critical:
            critical_fields = ", ".join(
                VALIDATION_FIELD_LABELS.get(issue.get("Field_Name", ""), issue.get("Field_Name", ""))
                for issue in critical
                if issue.get("Field_Name")
            )
            summary_parts.append(f"ขาดข้อมูล: {critical_fields}")
        if not summary_parts:
            summary_parts.append("ข้อมูล required ครบถ้วน")

        summaries[kpi_code] = {
            "status": status,
            "summary": "\n".join(summary_parts),
            "critical_fields": [issue.get("Field_Name", "") for issue in critical if issue.get("Field_Name")],
            "warning_fields": [],
        }
    return summaries


def generate_validation_report(output_path: Path, issues: list[dict[str, str]]) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Validation_Report"
    headers = ["Issue_Type", "KPI_Code", "Field_Name", "Message", "Sheet_Name"]
    sheet.append(headers)
    for issue in issues:
        sheet.append([issue.get(header, "") for header in headers])
    workbook.save(output_path)
    return output_path
